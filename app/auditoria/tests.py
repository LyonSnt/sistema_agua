from io import BytesIO

from django.contrib.auth.models import Group
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from usuarios.models import Usuario

from .models import Auditoria


class AuditoriaListaTests(TestCase):
    def crear_usuario(self, rol, username):
        grupo, _ = Group.objects.get_or_create(name=rol)
        usuario = Usuario.objects.create_user(
            username=username,
            password="clave-segura",
        )
        usuario.groups.add(grupo)
        return usuario

    def setUp(self):
        self.admin = self.crear_usuario("Administrador", "admin-auditoria")
        self.supervisor = self.crear_usuario("Supervisor", "supervisor-auditoria")

        Auditoria.objects.create(
            usuario=self.admin,
            accion="ANULAR_FACTURA",
            modulo="Facturación",
            descripcion="Anuló factura FAC001",
            objeto_repr="FAC001",
            ip="127.0.0.1",
        )
        Auditoria.objects.create(
            usuario=self.supervisor,
            accion="CAMBIAR_MEDIDOR",
            modulo="Medidores",
            descripcion="Cambió medidor MED001 por MED002",
            objeto_repr="MED002",
            ip="127.0.0.1",
        )

    def test_administrador_puede_ver_auditoria(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse("auditoria:lista"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Registro de acciones")
        self.assertContains(response, "Anuló factura FAC001")
        self.assertContains(response, "Cambió medidor MED001 por MED002")

    def test_supervisor_no_puede_ver_auditoria(self):
        self.client.force_login(self.supervisor)

        response = self.client.get(reverse("auditoria:lista"))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("panel:inicio"))

    def test_filtra_por_accion(self):
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse("auditoria:lista"),
            {"accion": "CAMBIAR_MEDIDOR"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cambió medidor MED001 por MED002")
        self.assertNotContains(response, "Anuló factura FAC001")

    def test_filtra_por_busqueda(self):
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse("auditoria:lista"),
            {"q": "FAC001"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Anuló factura FAC001")
        self.assertNotContains(response, "Cambió medidor MED001 por MED002")

    def test_administrador_exporta_excel_con_filtros(self):
        self.client.force_login(self.admin)

        response = self.client.get(
            reverse("auditoria:exportar_excel"),
            {"accion": "CAMBIAR_MEDIDOR"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        self.assertEqual(ws.title, "Auditoria")
        self.assertEqual(ws["A1"].value, "Fecha")
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws["C2"].value, "Cambiar medidor")
        self.assertEqual(ws["F2"].value, "Cambió medidor MED001 por MED002")
        self.assertTrue(
            Auditoria.objects.filter(
                accion="EXPORTAR_REPORTE",
                modulo="Auditoría",
                descripcion="Exportó registros de auditoría a Excel",
            ).exists()
        )

    def test_supervisor_no_puede_exportar_excel(self):
        self.client.force_login(self.supervisor)

        response = self.client.get(reverse("auditoria:exportar_excel"))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("panel:inicio"))
