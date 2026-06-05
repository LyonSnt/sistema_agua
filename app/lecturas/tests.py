from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from abonados.models import Abonado, Ruta, Sector
from medidores.models import Medidor
from usuarios.models import Usuario

from .models import Lectura, PeriodoFacturacion
from .views import (
    HOJA_CONTROL_LECTURAS,
    SESSION_IMPORTACION_CONTEXTO,
    SESSION_IMPORTACION_LECTURAS,
)
from . import views


class ImportarLecturasExcelTests(TestCase):
    def setUp(self):
        grupo = Group.objects.create(name="Lecturista")
        self.usuario = Usuario.objects.create_user(
            username="lecturista",
            password="clave-segura",
        )
        self.usuario.groups.add(grupo)
        self.client.force_login(self.usuario)

        self.sector = Sector.objects.create(nombre="Sector A")
        self.ruta = Ruta.objects.create(
            sector=self.sector,
            nombre="Ruta A",
        )
        self.otra_ruta = Ruta.objects.create(
            sector=self.sector,
            nombre="Ruta B",
        )

        self.abonado = Abonado.objects.create(
            codigo="AB001",
            cedula_ruc="0101010101",
            nombres="Ana",
            apellidos="Agua",
            direccion="Calle 1",
            sector=self.sector,
            ruta=self.ruta,
        )
        self.otro_abonado = Abonado.objects.create(
            codigo="AB002",
            cedula_ruc="0202020202",
            nombres="Bruno",
            apellidos="Bomba",
            direccion="Calle 2",
            sector=self.sector,
            ruta=self.otra_ruta,
        )

        self.medidor = Medidor.objects.create(
            abonado=self.abonado,
            numero="MED001",
            lectura_inicial=Decimal("10.00"),
        )
        self.otro_medidor = Medidor.objects.create(
            abonado=self.otro_abonado,
            numero="MED002",
            lectura_inicial=Decimal("20.00"),
        )

        self.periodo = PeriodoFacturacion.objects.create(
            nombre="Junio 2026",
            anio=2026,
            mes=6,
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 30),
            estado="ABIERTO",
        )

        self.lectura = Lectura.objects.create(
            periodo=self.periodo,
            medidor=self.medidor,
            lectura_anterior=Decimal("10.00"),
            lectura_actual=Decimal("10.00"),
            lectura_registrada=False,
        )
        self.lectura_fuera_de_ruta = Lectura.objects.create(
            periodo=self.periodo,
            medidor=self.otro_medidor,
            lectura_anterior=Decimal("20.00"),
            lectura_actual=Decimal("20.00"),
            lectura_registrada=False,
        )

    def descargar_workbook(self, **params):
        response = self.client.get(
            reverse("lecturas:plantilla_excel"),
            {
                "periodo": self.periodo.id,
                **params,
            },
        )
        self.assertEqual(response.status_code, 200)
        return load_workbook(BytesIO(response.content))

    def workbook_upload(self, wb, filename="lecturas.xlsx"):
        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        return SimpleUploadedFile(
            filename,
            archivo.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    def test_plantilla_excel_incluye_hoja_control_oculta_con_ids_permitidos(self):
        wb = self.descargar_workbook(ruta=self.ruta.id)

        self.assertIn("Lecturas", wb.sheetnames)
        self.assertIn(HOJA_CONTROL_LECTURAS, wb.sheetnames)

        control = wb[HOJA_CONTROL_LECTURAS]
        ids_visibles = [
            row[0]
            for row in wb["Lecturas"].iter_rows(
                min_row=2,
                max_col=1,
                values_only=True,
            )
            if row[0]
        ]
        ids_control = [
            row[0]
            for row in control.iter_rows(
                min_row=6,
                max_col=1,
                values_only=True,
            )
            if row[0]
        ]

        self.assertEqual(control.sheet_state, "hidden")
        self.assertEqual(control["B1"].value, self.periodo.id)
        self.assertEqual(control["B3"].value, str(self.ruta.id))
        self.assertEqual(ids_visibles, [self.lectura.id])
        self.assertEqual(ids_control, [self.lectura.id])

    def test_importacion_rechaza_id_que_no_pertenece_a_plantilla_original(self):
        wb = self.descargar_workbook(ruta=self.ruta.id)
        wb["Lecturas"]["A2"] = self.lectura_fuera_de_ruta.id
        wb["Lecturas"]["E2"] = "25.00"

        response = self.client.post(
            reverse("lecturas:importar_excel"),
            {"archivo": self.workbook_upload(wb, "manipulado.xlsx")},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "La lectura no pertenece a la plantilla original.",
        )

        session = self.client.session
        self.assertEqual(session[SESSION_IMPORTACION_LECTURAS], [])
        self.assertEqual(
            session[SESSION_IMPORTACION_CONTEXTO]["periodo_id"],
            self.periodo.id,
        )

    def test_confirmacion_actualiza_lectura_y_valida_contexto_de_periodo(self):
        wb = self.descargar_workbook()
        wb["Lecturas"]["E2"] = "18.50"

        response = self.client.post(
            reverse("lecturas:importar_excel"),
            {"archivo": self.workbook_upload(wb)},
        )
        self.assertEqual(response.status_code, 200)

        session = self.client.session
        self.assertEqual(len(session[SESSION_IMPORTACION_LECTURAS]), 1)
        self.assertEqual(
            session[SESSION_IMPORTACION_CONTEXTO]["periodo_id"],
            self.periodo.id,
        )

        response = self.client.post(
            reverse("lecturas:importar_excel"),
            {"confirmar": "1"},
        )

        self.assertRedirects(response, reverse("lecturas:importar_excel"))

        self.lectura.refresh_from_db()
        self.assertEqual(self.lectura.lectura_actual, Decimal("18.50"))
        self.assertTrue(self.lectura.lectura_registrada)

        session = self.client.session
        self.assertNotIn(SESSION_IMPORTACION_LECTURAS, session)
        self.assertNotIn(SESSION_IMPORTACION_CONTEXTO, session)

    def test_put_importar_lecturas_excel_no_esta_permitido(self):
        response = self.client.put(reverse("lecturas:importar_excel"))

        self.assertEqual(response.status_code, 405)

    def test_put_registro_masivo_lecturas_no_esta_permitido(self):
        response = self.client.put(reverse("lecturas:registro_masivo"))

        self.assertEqual(response.status_code, 405)

    def test_importacion_rechaza_extension_distinta_de_xlsx(self):
        archivo = SimpleUploadedFile(
            "lecturas.txt",
            b"no es excel",
            content_type="text/plain",
        )

        response = self.client.post(
            reverse("lecturas:importar_excel"),
            {"archivo": archivo},
            follow=True,
        )

        self.assertContains(response, "El archivo debe tener formato .xlsx.")
        self.assertNotIn(SESSION_IMPORTACION_LECTURAS, self.client.session)

    def test_importacion_rechaza_xlsx_corrupto_sin_error_500(self):
        archivo = SimpleUploadedFile(
            "lecturas.xlsx",
            b"contenido corrupto",
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response = self.client.post(
            reverse("lecturas:importar_excel"),
            {"archivo": archivo},
            follow=True,
        )

        self.assertContains(
            response,
            "No se pudo leer el archivo Excel. Descargue una nueva plantilla.",
        )
        self.assertNotIn(SESSION_IMPORTACION_LECTURAS, self.client.session)

    def test_importacion_rechaza_archivo_mayor_al_limite(self):
        limite_original = views.TAMANO_MAXIMO_EXCEL
        views.TAMANO_MAXIMO_EXCEL = 5

        try:
            archivo = SimpleUploadedFile(
                "lecturas.xlsx",
                b"123456",
                content_type=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

            response = self.client.post(
                reverse("lecturas:importar_excel"),
                {"archivo": archivo},
                follow=True,
            )
        finally:
            views.TAMANO_MAXIMO_EXCEL = limite_original

        self.assertContains(response, "El archivo no debe superar los 5 MB.")
        self.assertNotIn(SESSION_IMPORTACION_LECTURAS, self.client.session)
