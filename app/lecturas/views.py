from zipfile import BadZipFile

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render

from .models import PeriodoFacturacion
from .servicios import generar_lecturas_periodo

from django.db import transaction
from abonados.models import Sector, Ruta
from .models import Lectura
from decimal import Decimal
from usuarios.decoradores import rol_requerido

from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from auditoria.utils import registrar_auditoria

HOJA_CONTROL_LECTURAS = "_control"
SESSION_IMPORTACION_LECTURAS = "importacion_lecturas_validas"
SESSION_IMPORTACION_CONTEXTO = "importacion_lecturas_contexto"
TAMANO_MAXIMO_EXCEL = 5 * 1024 * 1024


def cargar_workbook_lecturas(archivo):
    nombre = (archivo.name or "").lower()

    if not nombre.endswith(".xlsx"):
        raise ValueError("El archivo debe tener formato .xlsx.")

    if archivo.size > TAMANO_MAXIMO_EXCEL:
        raise ValueError("El archivo no debe superar los 5 MB.")

    try:
        return load_workbook(archivo)
    except (InvalidFileException, BadZipFile, OSError, ValueError):
        raise ValueError(
            "No se pudo leer el archivo Excel. Descargue una nueva plantilla."
        )


@rol_requerido("Administrador", "Supervisor")
@require_http_methods(["GET", "POST"])
def generar_lecturas(request):
    periodos = PeriodoFacturacion.objects.filter(
        activo=True,
        estado="ABIERTO"
    ).order_by("-anio", "-mes")

    if request.method == "POST":
        periodo_id = request.POST.get("periodo")
        periodo = get_object_or_404(
            PeriodoFacturacion,
            id=periodo_id,
            activo=True,
            estado="ABIERTO",
        )

        resultado = generar_lecturas_periodo(periodo)

        registrar_auditoria(
            request,
            accion="LECTURA",
            modulo="Lecturas",
            descripcion=(
                f"Generó lecturas del período {periodo}. "
                f"Creadas: {resultado['creadas']}, "
                f"Existentes: {resultado['existentes']}, "
                f"Errores: {resultado['errores']}"
            ),
        )

        messages.success(
            request,
            f"Proceso finalizado. Lecturas creadas: {resultado['creadas']}. "
            f"Ya existentes: {resultado['existentes']}. "
            f"Errores: {resultado['errores']}."
        )

        return redirect("lecturas:generar")

    return render(request, "lecturas/generar.html", {
        "periodos": periodos,
    })

@rol_requerido("Administrador", "Supervisor", "Lecturista")
@require_http_methods(["GET", "POST"])
def registro_masivo_lecturas(request):
    periodos = PeriodoFacturacion.objects.filter(
        activo=True,
        estado="ABIERTO"
    ).order_by("-anio", "-mes")

    sectores = Sector.objects.filter(activo=True).order_by("nombre")
    rutas = Ruta.objects.filter(activo=True).select_related("sector").order_by("sector__nombre", "nombre")

    periodo_id = request.GET.get("periodo") or request.POST.get("periodo")
    sector_id = request.GET.get("sector") or request.POST.get("sector")
    ruta_id = request.GET.get("ruta") or request.POST.get("ruta")

    if sector_id in ["", "None", None]:
        sector_id = None

    if ruta_id in ["", "None", None]:
        ruta_id = None

    lecturas = Lectura.objects.none()
    periodo_seleccionado = None

    if periodo_id:
        periodo_seleccionado = get_object_or_404(
            PeriodoFacturacion,
            id=periodo_id,
            activo=True,
            estado="ABIERTO",
        )

        lecturas = Lectura.objects.select_related(
            "medidor",
            "medidor__abonado",
            "medidor__abonado__sector",
            "medidor__abonado__ruta",
        ).filter(
            periodo=periodo_seleccionado,
            activo=True,
        ).order_by(
            "medidor__abonado__sector__nombre",
            "medidor__abonado__ruta__nombre",
            "medidor__abonado__apellidos",
            "medidor__abonado__nombres",
        )

        if sector_id:
            lecturas = lecturas.filter(
                medidor__abonado__sector_id=sector_id
            )

        if ruta_id:
            lecturas = lecturas.filter(
                medidor__abonado__ruta_id=ruta_id
            )

    if request.method == "POST" and periodo_seleccionado:
        actualizadas = 0
        errores = 0

        with transaction.atomic():
            for lectura in lecturas:

                if hasattr(lectura, "factura"):
                    continue

                valor = request.POST.get(f"lectura_actual_{lectura.id}")

                if valor is None or valor == "":
                    continue

                try:
                    valor_decimal = Decimal(valor)

                    confirmar_sin_consumo = request.POST.get(
                        f"confirmar_sin_consumo_{lectura.id}"
                    )

                    if valor_decimal > lectura.lectura_anterior:
                        lectura.lectura_actual = valor_decimal
                        lectura.lectura_registrada = True

                    elif valor_decimal == lectura.lectura_anterior and confirmar_sin_consumo == "1":
                        lectura.lectura_actual = valor_decimal
                        lectura.lectura_registrada = True

                    else:
                        continue

                    lectura.actualizado_por = request.user
                    lectura.save()
                    actualizadas += 1

                except Exception as e:
                    errores += 1
                    print(f"Error lectura {lectura.id}: {e}")

        registrar_auditoria(
            request,
            accion="LECTURA",
            modulo="Lecturas",
            descripcion=(
                f"Registro masivo de lecturas. "
                f"Período ID: {periodo_id}, "
                f"Sector ID: {sector_id or 'Todos'}, "
                f"Ruta ID: {ruta_id or 'Todas'}, "
                f"Actualizadas: {actualizadas}, "
                f"Errores: {errores}."
            ),
        )

        messages.success(
            request,
            f"Lecturas actualizadas: {actualizadas}. Errores: {errores}."
        )

        return redirect(
            f"{request.path}?periodo={periodo_id}&sector={sector_id or ''}&ruta={ruta_id or ''}"
        )

    total_lecturas = lecturas.count()
    pendientes_lectura = 0
    listas_facturar = 0
    facturadas = 0

    for lectura in lecturas:
        if hasattr(lectura, "factura"):
            facturadas += 1
        elif not lectura.lectura_registrada:
            pendientes_lectura += 1
        else:
            listas_facturar += 1

    contexto = {
        "periodos": periodos,
        "sectores": sectores,
        "rutas": rutas,
        "lecturas": lecturas,
        "periodo_id": periodo_id,
        "sector_id": sector_id,
        "ruta_id": ruta_id,
        "contexto_periodo": {
            "total_lecturas": total_lecturas,
            "pendientes_lectura": pendientes_lectura,
            "listas_facturar": listas_facturar,
            "facturadas": facturadas,
        },
    }

    return render(request, "lecturas/registro_masivo.html", contexto)

@rol_requerido("Administrador", "Supervisor", "Lecturista")
def descargar_plantilla_lecturas(request):
    periodo_id = request.GET.get("periodo")
    sector_id = request.GET.get("sector")
    ruta_id = request.GET.get("ruta")

    periodo = get_object_or_404(PeriodoFacturacion, id=periodo_id, activo=True)

    lecturas = Lectura.objects.select_related(
        "medidor",
        "medidor__abonado",
        "medidor__abonado__sector",
        "medidor__abonado__ruta",
    ).filter(periodo=periodo, activo=True, lectura_registrada=False)

    if sector_id:
        lecturas = lecturas.filter(medidor__abonado__sector_id=sector_id)

    if ruta_id:
        lecturas = lecturas.filter(medidor__abonado__ruta_id=ruta_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas"

    ws.append(["ID Lectura", "Medidor", "Abonado", "Lectura anterior", "Lectura actual"])

    for lectura in lecturas:
        ws.append([
            lectura.id,
            lectura.medidor.numero,
            str(lectura.medidor.abonado),
            float(lectura.lectura_anterior),
            "",
        ])

    control = wb.create_sheet(HOJA_CONTROL_LECTURAS)
    control.sheet_state = "hidden"
    control.append(["periodo_id", periodo.id])
    control.append(["sector_id", sector_id or ""])
    control.append(["ruta_id", ruta_id or ""])
    control.append([])
    control.append(["lectura_id"])

    for lectura in lecturas:
        control.append([lectura.id])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="lecturas_{periodo.nombre}.xlsx"'
    wb.save(response)

    registrar_auditoria(
        request,
        accion="EXPORTAR_REPORTE",
        modulo="Lecturas",
        descripcion=f"Descargó plantilla de lecturas del período {periodo}",
        objeto=periodo,
    )

    return response

@rol_requerido("Administrador", "Supervisor", "Lecturista")
@require_http_methods(["GET", "POST"])
def importar_lecturas_excel(request):
    periodos = PeriodoFacturacion.objects.filter(activo=True, estado="ABIERTO")

    if request.method == "POST":
        confirmar = request.POST.get("confirmar")

        if confirmar == "1":
            filas_validas = request.session.get(SESSION_IMPORTACION_LECTURAS, [])
            contexto_importacion = request.session.get(SESSION_IMPORTACION_CONTEXTO)

            if not contexto_importacion:
                messages.error(
                    request,
                    "La sesión de importación expiró. Revise nuevamente el archivo."
                )
                return redirect("lecturas:importar_excel")

            periodo_id = contexto_importacion.get("periodo_id")
            ids_permitidos = set(
                contexto_importacion.get("lecturas_permitidas", [])
            )

            actualizadas = 0

            for item in filas_validas:
                lectura = Lectura.objects.get(
                    id=item["lectura_id"],
                    periodo_id=periodo_id,
                    activo=True,
                )

                if lectura.id not in ids_permitidos:
                    messages.error(
                        request,
                        "La importación contiene lecturas fuera de la plantilla original."
                    )
                    return redirect("lecturas:importar_excel")

                if hasattr(lectura, "factura"):
                    continue

                lectura.lectura_actual = Decimal(str(item["actual"]))
                lectura.lectura_registrada = True
                lectura.actualizado_por = request.user
                lectura.save()
                actualizadas += 1

            request.session.pop(SESSION_IMPORTACION_LECTURAS, None)
            request.session.pop(SESSION_IMPORTACION_CONTEXTO, None)

            registrar_auditoria(
                request,
                accion="IMPORTAR_LECTURAS",
                modulo="Lecturas",
                descripcion=(
                    f"Importó lecturas desde Excel. "
                    f"Lecturas actualizadas: {actualizadas}."
                ),
            )

            messages.success(
                request,
                f"Importación confirmada. Lecturas actualizadas: {actualizadas}."
            )

            return redirect("lecturas:importar_excel")

        archivo = request.FILES.get("archivo")

        if not archivo:
            messages.error(request, "Debe seleccionar un archivo Excel.")
            return redirect("lecturas:importar_excel")

        try:
            wb = cargar_workbook_lecturas(archivo)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("lecturas:importar_excel")

        ws = wb.active

        if HOJA_CONTROL_LECTURAS not in wb.sheetnames:
            messages.error(
                request,
                "El archivo no corresponde a una plantilla de lecturas vigente. Descargue una nueva plantilla."
            )
            return redirect("lecturas:importar_excel")

        control = wb[HOJA_CONTROL_LECTURAS]
        periodo_id = control["B1"].value
        sector_id = control["B2"].value or None
        ruta_id = control["B3"].value or None

        try:
            periodo = PeriodoFacturacion.objects.get(
                id=periodo_id,
                activo=True,
                estado="ABIERTO",
            )
        except PeriodoFacturacion.DoesNotExist:
            messages.error(
                request,
                "El período de la plantilla no está disponible para importación."
            )
            return redirect("lecturas:importar_excel")

        lecturas_permitidas = set()

        for fila in control.iter_rows(min_row=6, max_col=1, values_only=True):
            lectura_id_control = fila[0]

            if lectura_id_control:
                lecturas_permitidas.add(int(lectura_id_control))

        filas_validas = []
        filas_error = []

        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            lectura_id, medidor, abonado, lectura_anterior, lectura_actual = fila

            if not lectura_id or lectura_actual in [None, ""]:
                continue

            try:
                lectura = Lectura.objects.get(id=lectura_id, activo=True)

                if lectura.id not in lecturas_permitidas:
                    filas_error.append({
                        "fila": fila_num,
                        "medidor": medidor,
                        "abonado": abonado,
                        "error": "La lectura no pertenece a la plantilla original.",
                    })
                    continue

                if lectura.periodo_id != periodo.id:
                    filas_error.append({
                        "fila": fila_num,
                        "medidor": medidor,
                        "abonado": abonado,
                        "error": "La lectura no pertenece al período de la plantilla.",
                    })
                    continue

                if sector_id and lectura.medidor.abonado.sector_id != int(sector_id):
                    filas_error.append({
                        "fila": fila_num,
                        "medidor": medidor,
                        "abonado": abonado,
                        "error": "La lectura no pertenece al sector de la plantilla.",
                    })
                    continue

                if ruta_id and lectura.medidor.abonado.ruta_id != int(ruta_id):
                    filas_error.append({
                        "fila": fila_num,
                        "medidor": medidor,
                        "abonado": abonado,
                        "error": "La lectura no pertenece a la ruta de la plantilla.",
                    })
                    continue

                if hasattr(lectura, "factura"):
                    filas_error.append({
                        "fila": fila_num,
                        "medidor": medidor,
                        "abonado": abonado,
                        "error": "La lectura ya tiene factura generada.",
                    })
                    continue

                valor = Decimal(str(lectura_actual))

                if valor < lectura.lectura_anterior:
                    filas_error.append({
                        "fila": fila_num,
                        "medidor": medidor,
                        "abonado": abonado,
                        "error": "La lectura actual no puede ser menor que la anterior.",
                    })
                    continue

                filas_validas.append({
                    "lectura_id": lectura.id,
                    "periodo_id": lectura.periodo_id,
                    "medidor": medidor,
                    "abonado": abonado,
                    "anterior": str(lectura.lectura_anterior),
                    "actual": str(valor),
                    "consumo": str(valor - lectura.lectura_anterior),
                })

            except Exception as e:
                filas_error.append({
                    "fila": fila_num,
                    "medidor": medidor,
                    "abonado": abonado,
                    "error": str(e),
                })

        request.session[SESSION_IMPORTACION_LECTURAS] = filas_validas
        request.session[SESSION_IMPORTACION_CONTEXTO] = {
            "periodo_id": periodo.id,
            "sector_id": sector_id,
            "ruta_id": ruta_id,
            "lecturas_permitidas": list(lecturas_permitidas),
        }

        return render(request, "lecturas/importar_excel.html", {
            "periodos": periodos,
            "filas_validas": filas_validas,
            "filas_error": filas_error,
            "modo_preview": True,
            "periodo_importacion": periodo,
        })

    return render(request, "lecturas/importar_excel.html", {
        "periodos": periodos,
    })
