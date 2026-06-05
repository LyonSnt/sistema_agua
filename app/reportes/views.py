from decimal import Decimal

from django.contrib import messages
from django.shortcuts import render
from django.utils import timezone

from pagos.models import Pago
from facturacion.models import Factura
from usuarios.decoradores import rol_requerido
from django.http import HttpResponse
from openpyxl import Workbook
from datetime import date, datetime
from calendar import monthrange
from django.db.models.functions import TruncDate
from django.db.models import Count, Sum
from multas.models import Multa
from django.db.models import Prefetch
from django.template.loader import render_to_string
from weasyprint import HTML
from django.core.paginator import Paginator


def obtener_fecha_reporte(request, nombre_parametro="fecha"):
    valor = request.GET.get(nombre_parametro)

    if not valor:
        return timezone.localdate()

    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        messages.warning(
            request,
            "La fecha ingresada no es válida. Se usó la fecha actual."
        )
        return timezone.localdate()


def obtener_periodo_mensual_reporte(request):
    hoy = timezone.localdate()
    anio_param = request.GET.get("anio", hoy.year)
    mes_param = request.GET.get("mes", hoy.month)

    try:
        anio = int(anio_param)
        mes = int(mes_param)

        if mes < 1 or mes > 12:
            raise ValueError

        fecha_inicio = date(anio, mes, 1)
        fecha_fin = date(anio, mes, monthrange(anio, mes)[1])
    except (TypeError, ValueError):
        messages.warning(
            request,
            "El período ingresado no es válido. Se usó el mes actual."
        )
        anio = hoy.year
        mes = hoy.month
        fecha_inicio = date(anio, mes, 1)
        fecha_fin = date(anio, mes, monthrange(anio, mes)[1])

    return anio, mes, fecha_inicio, fecha_fin


@rol_requerido("Administrador", "Supervisor", "Cajero", "Consulta")
def cierre_diario(request):
    fecha = obtener_fecha_reporte(request)
    pagos = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "creado_por",
    ).prefetch_related(
        "factura__detalles"
    ).filter(
        fecha_pago__date=fecha,
        activo=True,
        anulado=False,
    )

    pagos_anulados = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "creado_por",
    ).filter(
        fecha_anulacion__date=fecha,
        activo=True,
        anulado=True,
    )

    total_recaudado = Decimal("0.00")
    total_agua = Decimal("0.00")
    total_alcantarillado = Decimal("0.00")
    total_multas = Decimal("0.00")
    total_otros = Decimal("0.00")

    for pago in pagos:
        total_recaudado += pago.valor_pagado

        factura = pago.factura

        if factura.total <= 0:
            continue

        proporcion = pago.valor_pagado / factura.total

        for detalle in factura.detalles.all():
            valor_proporcional = detalle.valor_total * proporcion

            if detalle.tipo == "AGUA":
                total_agua += valor_proporcional
            elif detalle.tipo == "ALCANTARILLADO":
                total_alcantarillado += valor_proporcional
            elif detalle.tipo == "MULTA":
                total_multas += valor_proporcional
            else:
                total_otros += valor_proporcional

    paginator = Paginator(pagos, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    paginator_anulados = Paginator(pagos_anulados, 10)
    page_anulados_number = request.GET.get("page_anulados")
    page_obj_anulados = paginator_anulados.get_page(page_anulados_number)


    contexto = {
        "fecha": fecha,
        "pagos": page_obj,
        "page_obj": page_obj,
        "querystring": f"fecha={fecha}",

        "total_pagos": pagos.count(),
        "total_recaudado": total_recaudado,
        "total_agua": total_agua,
        "total_alcantarillado": total_alcantarillado,
        "total_multas": total_multas,
        "total_otros": total_otros,

        "pagos_anulados": page_obj_anulados,
        "page_obj_anulados": page_obj_anulados,
        "total_pagos_anulados": pagos_anulados.count(),
    }

    return render(request, "reportes/cierre_diario.html", contexto)

@rol_requerido("Administrador", "Supervisor")
def cierre_diario_pdf(request):
    fecha = obtener_fecha_reporte(request)
    pagos = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "creado_por",
    ).prefetch_related(
        "factura__detalles"
    ).filter(
        fecha_pago__date=fecha,
        activo=True,
        anulado=False,
    )

    pagos_anulados = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "actualizado_por",
    ).filter(
        fecha_anulacion__date=fecha,
        activo=True,
        anulado=True,
    )

    total_recaudado = Decimal("0.00")
    total_agua = Decimal("0.00")
    total_alcantarillado = Decimal("0.00")
    total_multas = Decimal("0.00")
    total_otros = Decimal("0.00")

    for pago in pagos:
        total_recaudado += pago.valor_pagado
        factura = pago.factura

        if factura.total <= 0:
            continue

        proporcion = pago.valor_pagado / factura.total

        for detalle in factura.detalles.all():
            valor_proporcional = detalle.valor_total * proporcion

            if detalle.tipo == "AGUA":
                total_agua += valor_proporcional
            elif detalle.tipo == "ALCANTARILLADO":
                total_alcantarillado += valor_proporcional
            elif detalle.tipo == "MULTA":
                total_multas += valor_proporcional
            else:
                total_otros += valor_proporcional

    html_string = render_to_string(
        "reportes/cierre_diario_pdf.html",
        {
            "fecha": fecha,
            "pagos": pagos,
            "pagos_anulados": pagos_anulados,
            "total_pagos": pagos.count(),
            "total_recaudado": total_recaudado,
            "total_agua": total_agua,
            "total_alcantarillado": total_alcantarillado,
            "total_multas": total_multas,
            "total_otros": total_otros,
            "usuario": request.user,
        }
    )

    pdf = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="cierre_diario_{fecha}.pdf"'
    )
    return response

@rol_requerido("Administrador", "Supervisor", "Cajero", "Consulta")
def cartera_pendiente(request):
    busqueda = request.GET.get("q", "")

    facturas = Factura.objects.select_related(
        "abonado",
        "periodo",
    ).filter(
        estado__in=["PENDIENTE", "PARCIAL"],
        activo=True,
    ).order_by("abonado__apellidos", "abonado__nombres")

    if busqueda:
        facturas = (
            facturas.filter(numero__icontains=busqueda)
            | facturas.filter(abonado__nombres__icontains=busqueda)
            | facturas.filter(abonado__apellidos__icontains=busqueda)
            | facturas.filter(abonado__cedula_ruc__icontains=busqueda)
        )

    total_cartera = sum(f.saldo_pendiente for f in facturas)

    paginator = Paginator(facturas, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    contexto = {
        "facturas": page_obj,
        "page_obj": page_obj,
        "busqueda": busqueda,
        "querystring": f"q={busqueda}",
        "total_cartera": total_cartera,
    }

    return render(request, "reportes/cartera.html", contexto)

@rol_requerido("Administrador", "Supervisor", "Cajero", "Consulta")
def facturas_pagadas(request):
    busqueda = request.GET.get("q", "")

    pagos_validos = Pago.objects.filter(
        anulado=False
    ).select_related(
        "creado_por"
    ).order_by("-fecha_pago")

    facturas = Factura.objects.select_related(
        "abonado",
        "periodo",
    ).prefetch_related(
        Prefetch("pagos", queryset=pagos_validos, to_attr="pagos_validos")
    ).filter(
        estado="PAGADA",
        activo=True,
    ).order_by("-fecha_emision")

    if busqueda:
        facturas = (
            facturas.filter(numero__icontains=busqueda)
            | facturas.filter(abonado__nombres__icontains=busqueda)
            | facturas.filter(abonado__apellidos__icontains=busqueda)
            | facturas.filter(abonado__cedula_ruc__icontains=busqueda)
        )

    paginator = Paginator(facturas, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    contexto = {
        "facturas": page_obj,
        "page_obj": page_obj,
        "busqueda": busqueda,
        "querystring": f"q={busqueda}",
    }

    return render(request, "reportes/facturas_pagadas.html", contexto)

@rol_requerido("Administrador", "Supervisor", "Cajero", "Consulta")
def facturas_anuladas(request):
    busqueda = request.GET.get("q", "")

    facturas = Factura.objects.select_related(
        "abonado",
        "periodo",
        "actualizado_por",
    ).filter(
        estado="ANULADA",
        activo=True,
    ).order_by("-fecha_anulacion")

    if busqueda:
        facturas = (
            facturas.filter(numero__icontains=busqueda)
            | facturas.filter(abonado__nombres__icontains=busqueda)
            | facturas.filter(abonado__apellidos__icontains=busqueda)
            | facturas.filter(abonado__cedula_ruc__icontains=busqueda)
        )

    paginator = Paginator(facturas, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "reportes/facturas_anuladas.html", {
        "facturas": page_obj,
        "page_obj": page_obj,
        "busqueda": busqueda,
        "querystring": f"q={busqueda}",
    })

@rol_requerido("Administrador", "Supervisor", "Cajero", "Consulta")
def recaudacion_diaria(request):
    fecha = obtener_fecha_reporte(request)

    pagos = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "creado_por",
    ).prefetch_related(
        "factura__detalles"
    ).filter(
        fecha_pago__date=fecha,
        activo=True,
        anulado=False,
    )

    multas_cobradas = Multa.objects.select_related("abonado").filter(
        fecha_pago__date=fecha,
        activo=True,
        estado="PAGADA",
    )

    total_multas_administrativas = sum(
        multa.valor for multa in multas_cobradas
    )

    total_recaudado = Decimal("0.00")
    total_agua = Decimal("0.00")
    total_alcantarillado = Decimal("0.00")
    total_multas = Decimal("0.00")
    total_otros = Decimal("0.00")

    for pago in pagos:
        total_recaudado += pago.valor_pagado

        factura = pago.factura

        if factura.total <= 0:
            continue

        proporcion = pago.valor_pagado / factura.total

        for detalle in factura.detalles.all():
            valor = detalle.valor_total * proporcion

            if detalle.tipo == "AGUA":
                total_agua += valor
            elif detalle.tipo == "ALCANTARILLADO":
                total_alcantarillado += valor
            elif detalle.tipo == "MULTA":
                total_multas += valor
            else:
                total_otros += valor

    total_general = total_recaudado + total_multas_administrativas

    return render(request, "reportes/recaudacion_diaria.html", {
        "fecha": fecha,
        "pagos": pagos,
        "total_pagos": pagos.count(),
        "total_recaudado": total_recaudado,
        "total_agua": total_agua,
        "total_alcantarillado": total_alcantarillado,
        "total_multas": total_multas,
        "total_otros": total_otros,
        "multas_cobradas": multas_cobradas,
        "total_multas_administrativas": total_multas_administrativas,
        "total_general": total_general,
    })

@rol_requerido("Administrador", "Supervisor")
def exportar_recaudacion_diaria_excel(request):
    fecha = obtener_fecha_reporte(request)

    pagos = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "creado_por",
    ).filter(
        fecha_pago__date=fecha,
        activo=True,
        anulado=False,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Recaudación diaria"

    ws.append(["RECAUDACIÓN DIARIA"])
    ws.append(["Fecha", str(fecha)])
    ws.append([])

    ws.append([
        "Hora",
        "Factura",
        "Abonado",
        "Método",
        "Cajero",
        "Valor pagado",
    ])

    total = 0

    for pago in pagos:
        total += pago.valor_pagado

        ws.append([
            pago.fecha_pago.strftime("%H:%M"),
            pago.factura.numero,
            str(pago.factura.abonado),
            pago.metodo_pago,
            str(pago.creado_por or "-"),
            float(pago.valor_pagado),
        ])

    ws.append([])
    ws.append(["", "", "", "", "TOTAL", float(total)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="recaudacion_diaria_{fecha}.xlsx"'
    )

    wb.save(response)
    return response

@rol_requerido("Administrador", "Supervisor", "Cajero", "Consulta")
def recaudacion_mensual(request):
    anio, mes, fecha_inicio, fecha_fin = obtener_periodo_mensual_reporte(request)

    pagos = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "creado_por",
    ).prefetch_related(
        "factura__detalles"
    ).filter(
        fecha_pago__date__gte=fecha_inicio,
        fecha_pago__date__lte=fecha_fin,
        activo=True,
        anulado=False,
    )

    pagos_anulados = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "actualizado_por",
    ).filter(
        fecha_anulacion__date__gte=fecha_inicio,
        fecha_anulacion__date__lte=fecha_fin,
        activo=True,
        anulado=True,
    ).order_by("-fecha_anulacion")

    total_recaudado = Decimal("0.00")
    total_agua = Decimal("0.00")
    total_alcantarillado = Decimal("0.00")
    total_multas = Decimal("0.00")
    total_otros = Decimal("0.00")

    total_anulado = pagos_anulados.aggregate(
        total=Sum("valor_pagado")
    )["total"] or Decimal("0.00")

    for pago in pagos:
        total_recaudado += pago.valor_pagado

        factura = pago.factura
        if factura.total <= 0:
            continue

        proporcion = pago.valor_pagado / factura.total

        for detalle in factura.detalles.all():
            valor = detalle.valor_total * proporcion

            if detalle.tipo == "AGUA":
                total_agua += valor
            elif detalle.tipo == "ALCANTARILLADO":
                total_alcantarillado += valor
            elif detalle.tipo == "MULTA":
                total_multas += valor
            else:
                total_otros += valor

    resumen_diario = pagos.annotate(
        dia=TruncDate("fecha_pago")
    ).values("dia").annotate(
        total=Sum("valor_pagado"),
        cantidad=Count("id")
    ).order_by("dia")

    return render(request, "reportes/recaudacion_mensual.html", {
        "anio": anio,
        "mes": mes,
        "pagos": pagos,
        "resumen_diario": resumen_diario,
        "total_pagos": pagos.count(),
        "total_recaudado": total_recaudado,
        "total_agua": total_agua,
        "total_alcantarillado": total_alcantarillado,
        "total_multas": total_multas,
        "total_otros": total_otros,
        "pagos_anulados": pagos_anulados,
        "total_anulado": total_anulado,
    })

@rol_requerido("Administrador", "Supervisor")
def recaudacion_mensual_pdf(request):
    anio, mes, fecha_inicio, fecha_fin = obtener_periodo_mensual_reporte(request)

    pagos = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "creado_por",
    ).prefetch_related(
        "factura__detalles"
    ).filter(
        fecha_pago__date__gte=fecha_inicio,
        fecha_pago__date__lte=fecha_fin,
        activo=True,
        anulado=False,
    )

    total_recaudado = Decimal("0.00")
    total_agua = Decimal("0.00")
    total_alcantarillado = Decimal("0.00")
    total_multas = Decimal("0.00")
    total_otros = Decimal("0.00")

    for pago in pagos:
        total_recaudado += pago.valor_pagado

        factura = pago.factura

        if factura.total <= 0:
            continue

        proporcion = pago.valor_pagado / factura.total

        for detalle in factura.detalles.all():
            valor = detalle.valor_total * proporcion

            if detalle.tipo == "AGUA":
                total_agua += valor
            elif detalle.tipo == "ALCANTARILLADO":
                total_alcantarillado += valor
            elif detalle.tipo == "MULTA":
                total_multas += valor
            else:
                total_otros += valor

    resumen_diario = pagos.annotate(
        dia=TruncDate("fecha_pago")
    ).values("dia").annotate(
        total=Sum("valor_pagado"),
        cantidad=Count("id")
    ).order_by("dia")

    html_string = render_to_string(
        "reportes/recaudacion_mensual_pdf.html",
        {
            "anio": anio,
            "mes": mes,
            "pagos": pagos,
            "resumen_diario": resumen_diario,
            "total_pagos": pagos.count(),
            "total_recaudado": total_recaudado,
            "total_agua": total_agua,
            "total_alcantarillado": total_alcantarillado,
            "total_multas": total_multas,
            "total_otros": total_otros,
        }
    )

    pdf = HTML(string=html_string).write_pdf()

    response = HttpResponse(
        pdf,
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        f'inline; filename="recaudacion_mensual_{anio}_{mes}.pdf"'
    )

    return response

@rol_requerido("Administrador", "Supervisor")
def exportar_recaudacion_mensual_excel(request):
    anio, mes, fecha_inicio, fecha_fin = obtener_periodo_mensual_reporte(request)

    pagos = Pago.objects.select_related(
        "factura",
        "factura__abonado",
        "creado_por",
    ).filter(
        fecha_pago__date__gte=fecha_inicio,
        fecha_pago__date__lte=fecha_fin,
        activo=True,
        anulado=False,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Recaudación mensual"

    ws.append(["RECAUDACIÓN MENSUAL"])
    ws.append(["Año", anio])
    ws.append(["Mes", mes])
    ws.append([])

    ws.append([
        "Fecha",
        "Factura",
        "Abonado",
        "Método",
        "Cajero",
        "Valor pagado",
    ])

    total = 0

    for pago in pagos:
        total += pago.valor_pagado

        ws.append([
            pago.fecha_pago.strftime("%Y-%m-%d %H:%M"),
            pago.factura.numero,
            str(pago.factura.abonado),
            pago.metodo_pago,
            str(pago.creado_por or "-"),
            float(pago.valor_pagado),
        ])

    ws.append([])
    ws.append(["", "", "", "", "TOTAL", float(total)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="recaudacion_mensual_{anio}_{mes}.xlsx"'
    )

    wb.save(response)
    return response

@rol_requerido("Administrador", "Supervisor", "Cajero", "Consulta")
def cartera_vencida(request):

    facturas = Factura.objects.select_related(
        "abonado",
        "periodo",
    ).filter(
        activo=True,
        estado__in=["PENDIENTE", "PARCIAL"]
    ).order_by(
        "abonado__apellidos",
        "abonado__nombres",
        "periodo__anio",
        "periodo__mes",
    )

    resumen = {}

    for factura in facturas:
        abonado_id = factura.abonado.id

        if abonado_id not in resumen:
            resumen[abonado_id] = {
                "abonado": factura.abonado,
                "cantidad_facturas": 0,
                "total_deuda": Decimal("0.00"),
                "facturas": [],
            }

        resumen[abonado_id]["cantidad_facturas"] += 1
        resumen[abonado_id]["total_deuda"] += factura.saldo_pendiente
        resumen[abonado_id]["facturas"].append(factura)

    cartera = list(resumen.values())

    cartera.sort(
        key=lambda x: x["total_deuda"],
        reverse=True
    )

    total_deuda_general = sum(
        item["total_deuda"] for item in cartera
    )

    paginator = Paginator(cartera, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "reportes/cartera_vencida.html", {
        "cartera": page_obj,
        "page_obj": page_obj,
        "total_deuda_general": total_deuda_general,
        "total_abonados_mora": len(cartera),
    })

@rol_requerido("Administrador", "Supervisor")
def exportar_cartera_vencida_excel(request):

    facturas = Factura.objects.select_related(
        "abonado",
        "periodo",
    ).filter(
        activo=True,
        estado__in=["PENDIENTE", "PARCIAL"]
    ).order_by(
        "abonado__apellidos",
        "abonado__nombres",
    )

    resumen = {}

    for factura in facturas:
        abonado_id = factura.abonado.id

        if abonado_id not in resumen:
            resumen[abonado_id] = {
                "abonado": factura.abonado,
                "cantidad_facturas": 0,
                "total_deuda": Decimal("0.00"),
                "periodos": [],
            }

        resumen[abonado_id]["cantidad_facturas"] += 1
        resumen[abonado_id]["total_deuda"] += factura.saldo_pendiente

        resumen[abonado_id]["periodos"].append(
            f"{factura.periodo.nombre} (${factura.saldo_pendiente})"
        )

    cartera = list(resumen.values())

    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera vencida"

    ws.append(["CARTERA VENCIDA"])
    ws.append([])

    ws.append([
        "Abonado",
        "Cédula/RUC",
        "Facturas pendientes",
        "Total deuda",
        "Períodos pendientes",
    ])

    total_general = Decimal("0.00")

    for item in cartera:
        total_general += item["total_deuda"]

        ws.append([
            str(item["abonado"]),
            item["abonado"].cedula_ruc,
            item["cantidad_facturas"],
            float(item["total_deuda"]),
            ", ".join(item["periodos"]),
        ])

    ws.append([])
    ws.append(["", "", "", float(total_general)])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="cartera_vencida.xlsx"'
    )

    wb.save(response)

    return response


